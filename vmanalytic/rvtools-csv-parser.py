#!/usr/bin/env python3

import argparse
import csv
import os
import re
import sys

from openpyxl import Workbook


WHITELIST = ("vInfo", "vCPU", "vMemory", "vDisk", "vNetwork", "vHost")

REQUIRED_EXTRA_COLS = (
    "Environment",
    "Final OS",
    "Disk Size TB",
    "VM",
    "Cluster",
    "Disk Classification",
)


def get_csv_files(dir_path: str) -> list[str]:
    if not os.path.isdir(dir_path):
        return []
    return sorted(
        f for f in os.listdir(dir_path) if f.endswith(".csv")
    )


def normalize_xlsx_name(name: str) -> str:
    base = name.lower().strip()
    return base if base.endswith(".xlsx") else f"{base}.xlsx"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Merge RVTools CSV exports into one XLSX workbook."
    )
    parser.add_argument(
        "-p", "--path", help="Directory containing RVTools CSV files", required=True
    )
    parser.add_argument(
        "-o", "--outfile", help="Output XLSX filename (optional)", required=False
    )
    parser.add_argument(
        "-v",
        "--verbose",
        help="Verbose output",
        action="store_true",
        default=False,
    )
    args = parser.parse_args()

    data_dir = os.path.abspath(args.path)
    patt = r"_RVTools_export_all.*"
    instance = os.path.basename(re.sub(patt, "", data_dir)).lower()

    if args.outfile:
        outfile = normalize_xlsx_name(args.outfile)
    else:
        default_name = f"{instance}.xlsx"
        raw = input(
            f"Please give an output file to create [ Default: '{default_name}' ]: "
        ).strip()
        outfile = normalize_xlsx_name(raw) if raw else default_name

    files = get_csv_files(data_dir)
    if not files:
        print(f"No CSV files found in: {data_dir}", file=sys.stderr)
        return 1

    out_path = os.path.join(data_dir, outfile)
    if args.verbose:
        print(f"Writing spreadsheet '{out_path}'")

    wb = Workbook()
    del wb[wb.sheetnames[0]]

    for allowed in WHITELIST:
        for fname in files:
            if allowed not in fname:
                continue
            add_req_cols = {c: False for c in REQUIRED_EXTRA_COLS}
            rows = 0
            if args.verbose:
                print(f"Adding tab '{allowed}'")
            try:
                csv_path = os.path.join(data_dir, fname)
                with open(csv_path, newline="", encoding="utf8", errors="ignore") as f_input:
                    ws = wb.create_sheet(title=allowed)
                    for row in csv.reader(f_input, delimiter=","):
                        if rows == 0:
                            if ("env" in row) and allowed in ("vInfo", "vDisk"):
                                if args.verbose:
                                    print(
                                        "Converting the env column to Environment column"
                                    )
                                row = ["Environment" if x == "env" else x for x in row]
                            for col in add_req_cols:
                                if col not in row and allowed == "vInfo":
                                    row.append(col)
                                    add_req_cols[col] = True
                                    if args.verbose:
                                        print(f"Adding Column Header: {col}")
                        else:
                            for col in add_req_cols:
                                if add_req_cols[col] and allowed == "vInfo":
                                    row.append("")
                        ws.append(row)
                        rows += 1
                if args.verbose:
                    print(f"Successfully wrote tab '{allowed}'")
            except OSError as e:
                print(
                    f"Error writing tab '{allowed}': {e}",
                    file=sys.stderr,
                )
                return 1

    try:
        wb.save(out_path)
        if args.verbose:
            print(f"Successfully wrote '{out_path}'")
    except OSError as e:
        print(f"Error saving '{out_path}': {e}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
