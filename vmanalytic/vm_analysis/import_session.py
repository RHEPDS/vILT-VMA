"""
Prepare a temporary project directory for run_analysis():
- data/index.xlsx + RVTools .xlsx file(s)
- helper_files/ (copy from repo)
- saved_csv_files/
"""

from __future__ import annotations

import re
import shutil
import uuid
from pathlib import Path

import pandas as pd

# Parent folder for all import sessions (under repo)
UPLOADS_SUBDIR = Path("data") / "uploads"


def _slug(name: str) -> str:
    s = name.strip().lower()
    s = re.sub(r"[^a-z0-9._-]+", "_", s)
    return s or "vcenter"


def copy_helper_files(repo_root: Path, dest_root: Path) -> None:
    src = repo_root / "helper_files"
    dst = dest_root / "helper_files"
    if dst.exists():
        shutil.rmtree(dst)
    shutil.copytree(src, dst)


def normalize_rvtools_sheet_names(path: Path) -> None:
    """
    Rename worksheets to the exact names expected by the analysis (vInfo, vHost, vDisk).
    RVTools / Excel often differ only by case (e.g. vinfo vs vInfo).
    """
    import openpyxl

    if path.suffix.lower() == ".xls":
        raise ValueError(
            "Legacy .xls format is not supported. In Excel: Save As → Excel Workbook (.xlsx)."
        )

    wb = openpyxl.load_workbook(path, read_only=False, keep_vba=True)
    try:
        by_lower = {s.lower(): s for s in wb.sheetnames}
        canonical = [("vinfo", "vInfo"), ("vhost", "vHost"), ("vdisk", "vDisk")]
        # Pass 1: move mismatched names to temporary titles (avoid collisions)
        tmp_renames: list[tuple[str, str]] = []
        for i, (low, proper) in enumerate(canonical):
            if low not in by_lower:
                continue
            old = by_lower[low]
            if old == proper:
                continue
            tmp = f"__rvnorm_{i}_{low}__"
            wb[old].title = tmp
            tmp_renames.append((tmp, proper))
        # Pass 2: final canonical names
        for tmp, proper in tmp_renames:
            wb[tmp].title = proper
        wb.save(path)
    finally:
        wb.close()


def write_index_xlsx(path: Path, vcenter_names: list[str]) -> None:
    """Minimal index workbook for the analysis pipeline."""
    df = pd.DataFrame(
        {
            "vCenter": vcenter_names,
            "In Scope": [True] * len(vcenter_names),
        }
    )
    df.to_excel(path, sheet_name="index", index=False)


def session_from_uploaded_xlsx(
    repo_root: Path,
    saved_paths: list[tuple[str, Path]],
) -> tuple[Path, int]:
    """
    saved_paths: list of (original_filename, temp_path_to_uploaded_bytes)
    Produces one .xlsx per file under session data/, named so vCenter string matches.
    """
    session_id = str(uuid.uuid4())
    session_root = (repo_root / UPLOADS_SUBDIR / session_id).resolve()
    data_dir = session_root / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    (session_root / "saved_csv_files").mkdir(parents=True, exist_ok=True)
    copy_helper_files(repo_root, session_root)

    vcenters: list[str] = []
    seen: set[str] = set()
    for i, (orig_name, tmp_path) in enumerate(saved_paths):
        stem = Path(orig_name).stem
        vc = _slug(stem)
        if vc in seen:
            vc = f"{vc}_{i}"
        seen.add(vc)
        vcenters.append(vc)
        dest_name = f"{vc}.xlsx"
        shutil.copy2(tmp_path, data_dir / dest_name)

    index_path = data_dir / "index.xlsx"
    write_index_xlsx(index_path, vcenters)
    # Default index rows (19) with headroom for larger inventories
    index_nrows = min(500, max(19, len(vcenters) + 2))
    return session_root, index_nrows


def validate_rvtools_xlsx(path: Path) -> None:
    """Ensure required worksheets exist (after normalize_rvtools_sheet_names)."""
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("openpyxl is required to validate uploads") from e
    wb = None
    try:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            wb = openpyxl.load_workbook(path, read_only=False, keep_vba=True)
        names = set(wb.sheetnames)
    finally:
        if wb is not None:
            wb.close()
    missing = {"vInfo", "vHost", "vDisk"} - names
    if missing:
        raise ValueError(
            f"Missing worksheet(s): {', '.join(sorted(missing))}. "
            f"Found: {', '.join(sorted(names))}"
        )


def session_from_vcenter_xlsx(repo_root: Path, xlsx_path: Path, vc_label: str) -> tuple[Path, int]:
    """Place a single RVTools-like export from vCenter collection."""
    session_id = str(uuid.uuid4())
    session_root = (repo_root / UPLOADS_SUBDIR / session_id).resolve()
    data_dir = session_root / "data"
    data_dir.mkdir(parents=True, exist_ok=True)
    (session_root / "saved_csv_files").mkdir(parents=True, exist_ok=True)
    copy_helper_files(repo_root, session_root)

    vc = _slug(vc_label)
    shutil.copy2(xlsx_path, data_dir / f"{vc}.xlsx")
    write_index_xlsx(data_dir / "index.xlsx", [vc])
    return session_root, 19
